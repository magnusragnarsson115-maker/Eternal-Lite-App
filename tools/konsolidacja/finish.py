# -*- coding: utf-8 -*-
"""Macierz funkcji do XLSX + wstrzykniecie rozstrzygniec do obu roadmap HTML."""
import json
import html
import os

E = html.escape
ROWS = json.load(open('build/MACIERZ.json'))

MONET = [
 ('K0', 'Aplikacja pacjenta — DARMOWA', 'Zero opłat dla pacjenta. Warunek skali i jakości zbioru danych.'),
 ('K1', 'Subskrypcje niepacjenckie', 'Pet, Vault/Legacy, immersja premium — poza rdzeniem pacjenta.'),
 ('K2', 'Hardware i wkłady', 'Station: zakup 1 499 PLN lub HaaS 249 PLN/mies; wkłady 149 PLN/mies.'),
 ('K3', 'API i eksport danych', 'Płatny dostęp programistyczny; dane wyłącznie zagregowane i zanonimizowane.'),
 ('K4', 'Eternal Token i Forge', 'Gospodarka wewnętrzna marketplace modułów i IP.'),
 ('K5', 'Prowizje marketplace', 'Telemedycyna 20–30%, laboratoria 5–15%, apteka i suplementy.'),
 ('K6', 'Płatnicy i ubezpieczyciele', 'Scoring B2B, składka pay-as-you-live, programy prewencyjne.'),
 ('K7', 'Przychodnie i lekarze', 'Eternal Assist (AI Scribe) 99–199 PLN/mies za gabinet; PUPM 15–25 PLN.'),
 ('K8', 'Granty i dotacje', 'NCBR do 500 tys. bez wkładu własnego, PARP, FENG, Horizon Europe.'),
 ('K9', 'Licencjonowanie IP', 'Royalty 5–15% z Fundacji do spółki; white-label dla partnerów.'),
 ('K10', 'Fitness i wellness', 'Plany treningowe, suplementacja, Auto-Refill, corporate wellness.'),
 ('K11', 'Choroby przewlekłe', 'Pakiety dla diabetyków, kardiologii i zdrowia psychicznego — B2B klinika.'),
]

# ---------- XLSX ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'Macierz funkcji'
HEAD = ['Kod', 'Nazwa funkcji', 'Produkt', 'Moduł', 'Etap', 'Klasa MDR',
        'Czy zarabia', 'Kanał monetyzacji', 'Potrzeba użytkownika', 'Potrzeba ekosystemu',
        'Duplikacja w efekcie', 'Dubluje się z', 'Na czym polega duplikacja',
        'Liczba źródeł', 'Pliki źródłowe']
ws.append(HEAD)
for c in range(1, len(HEAD) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F3864')
    cell.alignment = Alignment(vertical='center', wrap_text=True)
for r in ROWS:
    ws.append([r['kod'], r['nazwa'], r['produkt'], r['modul'], r['etap'], r['klasa'],
               r['zarabia'], r['kanal'], r['uzytkownik'], r['ekosystem'],
               r['dup_efekt'], r['dup_z'], r['dup_opis'], r['n_zrodel'], r['zrodla']])
for i, w in enumerate([9, 42, 20, 34, 8, 10, 15, 46, 18, 18, 28, 20, 52, 11, 40], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

ws2 = wb.create_sheet('Model monetyzacji')
ws2.append(['Kanał', 'Nazwa', 'Istota'])
for c in range(1, 4):
    cell = ws2.cell(row=1, column=c)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F3864')
for m in MONET:
    ws2.append(list(m))
for i, w in enumerate([9, 36, 86], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

out_x = '/home/user/Eternal-Lite-App/out/ETERNAL_MACIERZ_FUNKCJI.xlsx'
wb.save(out_x)
print(out_x, os.path.getsize(out_x), 'B,', len(ROWS), 'funkcji')

# ---------- sekcja HTML do wstrzykniecia ----------
mon_rows = ''.join('<tr><td><b>%s</b></td><td>%s</td><td>%s</td></tr>'
                   % (E(a), E(b), E(c)) for a, b, c in MONET)
dupg = {}
for r in ROWS:
    if r['dup_efekt']:
        dupg.setdefault(r['dup_efekt'], (r['dup_opis'], []))[1].append(r['kod'])
dup_rows = ''.join('<tr><td>%s</td><td>%s</td><td>%s</td></tr>'
                   % (E(k), E(', '.join(sorted(v[1]))), E(v[0])) for k, v in dupg.items())
zar = sum(1 for r in ROWS if r['zarabia'] == 'TAK')

SEK = """
<section id="rozstrzygniecia" style="background:#fffdf5;border-top:3px solid #e0a33e">
<h2>Rozstrzygnięcia z biznesplanu i specyfikacji</h2>
<p class="lead">Ta sekcja jest nowa: przenosi do roadmapy decyzje wypracowane w dokumentach
biznesowym i technicznym. <b>Etapy 7&ndash;11 pozostają bez zmian</b> &mdash; w źródłach są jawnie
oznaczone jako fikcja i worldbuilding, więc nie podlegają rozstrzygnięciom biznesowym.</p>

<h3>Model monetyzacji &mdash; wersja przyjęta</h3>
<p class="lead">Rozstrzygnięcie sporu cenowego: <b>aplikacja pacjenta jest darmowa w całości</b>
(za Master 5.4, wersją najnowszą). Przychód przenosi się na jedenaście kanałów wokół niej.
Odrzucone warianty: 29,99/49,99 PLN z oficjalnego decku, 49 PLN z checklist, 19&ndash;29 PLN z planu
operacyjnego &mdash; wszystkie pochodzą z dokumentów starszych niż Master 5.4.</p>
<table><thead><tr><th>Kanał</th><th>Nazwa</th><th>Istota</th></tr></thead><tbody>%s</tbody></table>

<h3>Macierz funkcji &mdash; %d funkcji w 43 modułach</h3>
<p class="lead">Pełna macierz w pliku <code>ETERNAL_MACIERZ_FUNKCJI.xlsx</code> oraz w specyfikacji.
Z %d funkcji <b>%d ma przypisany kanał monetyzacji</b>, a <b>%d to funkcje fundamentowe</b>,
które nie zarabiają wprost, ale bez nich reszta nie działa.</p>

<h3>Duplikacja w efekcie końcowym</h3>
<p class="lead">Nie chodzi o duplikat mechanizmu, tylko o ten sam rezultat dla użytkownika
osiągany różnymi drogami. %d funkcji w %d grupach. To kandydaci do scalenia lub do świadomej
decyzji, że redundancja jest celowa (np. alert ratunkowy ma trzy drogi umyślnie).</p>
<table><thead><tr><th>Grupa</th><th>Funkcje</th><th>Na czym polega</th></tr></thead><tbody>%s</tbody></table>

<h3>Eternal Forge &mdash; rozstrzygnięcie sprzeczności</h3>
<p class="lead">Źródła nie były zgodne. Checklista <i>enriched</i> opisuje Forge jako warstwę
produkcji hardware; Macierz 40 Projektów i wszystkie wersje v3&ndash;v5 &mdash; jako marketplace IP
i API z gospodarką tokenową. <b>Przyjęta wersja: marketplace IP/API</b>, bo jest spójna z resztą
korpusu i z modułem A16. Warstwa produkcji hardware należy do Eternal Station, nie do Forge.</p>
</section>
""" % (mon_rows, len(ROWS), len(ROWS), zar, len(ROWS) - zar,
       sum(1 for r in ROWS if r['dup_efekt']), len(dupg), dup_rows)

for f in ['ETERNAL_ROADMAPA_SCALONA.html', 'ETERNAL_ROADMAPA_APLIKACJA.html']:
    p = '/home/user/Eternal-Lite-App/out/' + f
    s = open(p, encoding='utf-8').read()
    if 'id="rozstrzygniecia"' in s:
        continue
    anchor = '<footer>'
    s = s.replace(anchor, SEK + anchor, 1)
    s = s.replace('<a href="#zrodla">Indeks', '<a href="#rozstrzygniecia">Rozstrzygnięcia</a><a href="#zrodla">Indeks', 1)
    open(p, 'w', encoding='utf-8').write(s)
    print('zaktualizowano', f, os.path.getsize(p), 'B')
