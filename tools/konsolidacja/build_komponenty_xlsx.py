# -*- coding: utf-8 -*-
"""Macierz komponentow: 337 funkcji x komponenty, dostawcy, warstwa, certyfikacja."""
import json
import os
import sys
import collections

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dane_komponenty import (K, WARSTWA, SZCZEBEL, WYZWALACZE, SKLADOWE,
                             EKONOMIA, MODULY, BEZPIECZENSTWO)
from dane_moduly import M as MOD, KUBELKI

R = json.load(open('build/KOMPONENTY.json'))
OUT = '/home/user/Eternal-Lite-App/out/ETERNAL_MACIERZ_KOMPONENTOW.xlsx'

GRAN = PatternFill('solid', fgColor='1B3A6B')
RDZA = PatternFill('solid', fgColor='B8431F')
TLA = {'A': PatternFill('solid', fgColor='D7F0DD'),
       'B': PatternFill('solid', fgColor='F3E3C3'),
       'C': PatternFill('solid', fgColor='F8D7DA')}
CIENKA = Border(*[Side('thin', color='D8D4CE')] * 4)
wb = Workbook()


def arkusz(nazwa, naglowki, wiersze, szer, kolor_kol=None, zawijaj=None, naglowek_fill=GRAN):
    ws = wb.create_sheet(nazwa)
    ws.append(naglowki)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = naglowek_fill
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 32
    for w in wiersze:
        ws.append(w)
    for i, s in enumerate(szer, 1):
        ws.column_dimensions[get_column_letter(i)].width = s
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical='top',
                                    wrap_text=(zawijaj is None or c.column in zawijaj))
            c.font = Font(size=9)
            c.border = CIENKA
        if kolor_kol:
            v = str(row[kolor_kol - 1].value or '')
            if v in TLA:
                row[kolor_kol - 1].fill = TLA[v]
                row[kolor_kol - 1].font = Font(size=9, bold=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    return ws


# --- 1. Funkcje -> komponenty ---------------------------------------------
arkusz('Funkcje-komponenty',
 ['Kod', 'Funkcja', 'Produkt', 'Moduł', 'Klasa komponentu', 'Klasy wspierające',
  'Składowe I–IV', 'Dostawca na start', 'Warstwa', 'Charakter', 'Dlaczego ta warstwa',
  'Granica', 'Etap', 'Certyfikacja', 'Co konkretnie', 'Szczebel kontroli',
  'Czas wyjścia (dni)', 'Próg zmiany modelu'],
 [[r['kod'], r['nazwa'], r['produkt'], '%s — %s' % (r['modul'], r['modul_nazwa']),
   '%s — %s' % (r['klasa'], r['klasa_nazwa']), r['wspierajace'], r['skladowe'],
   r['dostawca_start'], r['warstwa'], r['charakter'], r['warstwa_powod'], r['granica'],
   r['etap'], r['certyfikacja'], r['certyfikacja_co'],
   '%d — %s' % (r['szczebel'], r['szczebel_nazwa']), r['czas_wyjscia'], r['prog_zmiany']]
  for r in R],
 [8, 44, 17, 30, 34, 14, 30, 40, 6, 11, 40, 8, 8, 26, 40, 26, 9, 40],
 kolor_kol=9)

# --- 2. Klasy komponentow --------------------------------------------------
lic = collections.Counter(r['klasa'] for r in R)
lic_w = collections.Counter()
for r in R:
    lic_w[r['klasa']] += 1
    for x in r['wspierajace'].split(' + '):
        if x:
            lic_w[x] += 1
arkusz('Klasy komponentów',
 ['Klasa', 'Nazwa', 'Funkcji wiodących', 'Funkcji łącznie', 'Składowe wymagane',
  '(A) Open source / darmowe', '(B) Płatne / licencjonowane', '(C) Własne / docelowe',
  'Próg wyjścia', 'Mechanizm kontroli i ryzyko', 'Rekomendacja na start',
  'Szczebel docelowy', 'Źródło'],
 [[kod, v[0], lic.get(kod, 0), lic_w.get(kod, 0),
   ' / '.join('%s. %s' % (s, SKLADOWE[s][0]) for s in v[8]),
   v[1], v[2], v[3], v[4], v[5], v[6],
   '%d — %s' % (v[7], SZCZEBEL[v[7]][0]), v[9]]
  for kod, v in sorted(K.items())],
 [7, 34, 10, 10, 30, 46, 46, 40, 30, 50, 40, 26, 34])

# --- 3. Brama 33% ----------------------------------------------------------
arkusz('Brama 33% — ekonomia',
 ['Klasa', 'Dostawca', 'Model rozliczenia', 'Koszt PLN / user / mies',
  'Udział docelowy %', 'Podstawa liczby'],
 [[k, d, m, c, u, p] for k, d, m, c, u, p in EKONOMIA],
 [7, 38, 34, 14, 12, 56])
ws = wb['Brama 33% — ekonomia']
for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
    for c in row:
        if isinstance(c.value, int) and c.value > 33:
            c.fill = TLA['C']
            c.font = Font(size=9, bold=True)

# --- 4. Wyzwalacze zmiany modelu ------------------------------------------
arkusz('Kiedy zmienić model',
 ['Kod', 'Wyzwalacz', 'Co konkretnie jest mierzone', 'Co się zmienia', 'Źródło'],
 [list(w) for w in WYZWALACZE], [6, 34, 60, 60, 40], naglowek_fill=RDZA)

# --- 5. Strategia wobec gotowych modulow ----------------------------------
arkusz('Gotowe moduły — strategia',
 ['Podmiot', 'Co robi', 'Co daje nam', 'Czego nie daje', 'Nasza postawa',
  'Warunek zmiany postawy'],
 [list(m) for m in MODULY], [34, 46, 46, 46, 46, 52])

# --- 6. Bezpieczenstwo -----------------------------------------------------
arkusz('Bezpieczeństwo bramy',
 ['Zasada', 'Na czym polega', 'Źródło'],
 [list(b) for b in BEZPIECZENSTWO], [40, 76, 44])

# --- 7. Certyfikacja -------------------------------------------------------
cert = [r for r in R if r['warstwa'] == 'C']
mod_c = collections.Counter('%s — %s' % (r['modul'], r['modul_nazwa']) for r in cert)
arkusz('Certyfikacja — zakres',
 ['Moduł', 'Funkcji w warstwie C', 'Etapy', 'Klasa komponentu wiodąca', 'Ścieżka'],
 [[m, n,
   ', '.join(sorted({r['etap'] for r in cert if '%s — %s' % (r['modul'], r['modul_nazwa']) == m})),
   ', '.join(sorted({r['klasa'] for r in cert if '%s — %s' % (r['modul'], r['modul_nazwa']) == m})),
   ('Proxy do cudzego CE (Labplus) albo własne dossier'
    if any(r['klasa'] == 'K28' for r in cert
           if '%s — %s' % (r['modul'], r['modul_nazwa']) == m)
    else 'Dossier klasy IIa: 80–150 tys. zł, 6–12 mies.')]
  for m, n in mod_c.most_common()],
 [40, 12, 22, 22, 50], naglowek_fill=RDZA)

# --- 8. Modul jako calosc: kupic / zarzadzac / budowac --------------------
FUN = collections.Counter(r['modul'] for r in R)


def postawa(pokr):
    if pokr >= 80:
        return 'AGREGUJEMY — kandydat na cały moduł'
    if pokr >= 30:
        return 'ZARZĄDZAMY — kandydat częściowy, moduł składamy sami'
    return 'BUDUJEMY — brak kandydata rynkowego'


arkusz('Moduł jako całość',
 ['Moduł', 'Nazwa', 'Funkcji', 'Postawa', 'Kandydat na cały moduł', 'Pokrycie %',
  'Kontrola %', 'Co zostaje nasze bezwzględnie', 'Alternatywa open source',
  'Kiedy budujemy własne', 'Adapter', 'Uzasadnienie adaptera', 'Priorytet', 'Owner'],
 [[k, v[0], str(FUN.get(k, 0)), postawa(v[7]), v[6], v[7], v[8], v[9], v[10], v[11],
   v[12][0], v[12][1], v[14], v[15]]
  for k, v in sorted(MOD.items(),
                     key=lambda x: (x[0][0], int(x[0][1:])))],
 [7, 34, 8, 34, 46, 9, 9, 40, 46, 40, 8, 40, 8, 12])
ws = wb['Moduł jako całość']
for row in ws.iter_rows(min_row=2):
    v = str(row[3].value or '')
    row[3].fill = (TLA['A'] if v.startswith('AGREG') else
                   TLA['B'] if v.startswith('ZARZ') else TLA['C'])
    row[3].font = Font(size=9, bold=True)

# --- 9. Wellness -> medyczne: cztery kubelki ------------------------------
arkusz('Wellness → medyczne',
 ['Moduł', 'Nazwa', 'Kubełek', 'Co znaczy', 'Co z tym robić', 'Funkcji',
  'Warstwa A', 'Warstwa B', 'Warstwa C'],
 [[k, v[0], v[13], KUBELKI[v[13]][0] + ' — ' + KUBELKI[v[13]][1], KUBELKI[v[13]][2],
   str(FUN.get(k, 0)),
   str(sum(1 for r in R if r['modul'] == k and r['warstwa'] == 'A')),
   str(sum(1 for r in R if r['modul'] == k and r['warstwa'] == 'B')),
   str(sum(1 for r in R if r['modul'] == k and r['warstwa'] == 'C'))]
  for k, v in sorted(MOD.items(), key=lambda x: (x[0][0], int(x[0][1:])))],
 [7, 34, 8, 60, 50, 8, 9, 9, 9], naglowek_fill=RDZA)
ws = wb['Wellness → medyczne']
KOL = {'W': 'D7F0DD', 'W>M': 'E3EDFA', 'M>W': 'F3E3C3', 'M': 'F8D7DA'}
for row in ws.iter_rows(min_row=2):
    v = str(row[2].value or '')
    if v in KOL:
        row[2].fill = PatternFill('solid', fgColor=KOL[v])
        row[2].font = Font(size=9, bold=True)

# --- 8. Warstwy i szczeble (legenda) --------------------------------------
ws = wb.create_sheet('Legenda', 0)
ws.append(['ETERNAL LIFE — MACIERZ KOMPONENTÓW'])
ws['A1'].font = Font(bold=True, size=16, color='B8431F')
c_w = collections.Counter(r['warstwa'] for r in R)
info = [
 [],
 ['Co zawiera ten skoroszyt'],
 ['Funkcje-komponenty', 'Każda z %d funkcji z przypisaną klasą komponentu, dostawcą, '
  'warstwą zgodności, etapem, wymogiem certyfikacji, czasem wyjścia i progiem zmiany '
  'modelu. Realizuje polecenie zapisane w Master 5.4: „Do każdej karty funkcji dopisać: '
  'KLASA KOMPONENTU, WARSTWA ZGODNOŚCI (A/B/C), CZAS WYJŚCIA w dniach i PRÓG ZMIANY '
  'wyrażony liczbą”.' % len(R)],
 ['Klasy komponentów', '%d klas z trzema wariantami każda (A open source / B płatny / '
  'C własny), progiem wyjścia i mechanizmem kontroli. 25 klas pochodzi z Master 5.4; '
  'K10, K19, K24, K29 i K30 uzupełniają lukę.' % len(K)],
 ['Brama 33% — ekonomia', 'Koszt każdego wariantu w przeliczeniu na użytkownika '
  'miesięcznie i docelowy udział w klasie. Reguła 33%: żaden dostawca nie może '
  'obsługiwać więcej niż jednej trzeciej aktywnych użytkowników w obrębie klasy.'],
 ['Kiedy zmienić model', 'Osiem wyzwalaczy, po których dotychczasowy model przestaje '
  'obowiązywać — od progu kosztowego po umieszczenie własnego logo na cudzym wyrobie.'],
 ['Gotowe moduły — strategia', 'Vitalera, Labplus, twojpsycholog.ai, sieci laboratoryjne, '
  'agregatory wearables: co dają, czego nie dają, jaka postawa i kiedy ją zmienić.'],
 ['Bezpieczeństwo bramy', 'Osiem zasad, bez których brama jest tylko routerem.'],
 ['Certyfikacja — zakres', 'Które moduły wymagają certyfikacji i którą ścieżką.'],
 ['Moduł jako całość', 'Dla każdego z 43 modułów: czy istnieje kandydat na cały moduł, '
  'jakie ma pokrycie, ile kontroli zostawia i co mimo wszystko zostaje nasze. '
  'Postawa AGREGUJEMY / ZARZĄDZAMY / BUDUJEMY wynika z pokrycia, nie z preferencji.'],
 ['Wellness → medyczne', 'Cztery kubełki: wellness na zawsze, wellness dziś a medyczna '
  'docelowo, medyczna z natury wydana jako wellness (najwyższe ryzyko), medyczna od '
  'początku.'],
 [],
 ['Warstwy zgodności'],
 ['Warstwa', 'Zakres', 'Charakter', 'Certyfikacja', 'Funkcji', 'Co to znaczy w praktyce'],
]
for w, v in WARSTWA.items():
    info.append([w, v[0], v[2], v[3], c_w.get(w, 0), v[1]])
info += [
 [],
 ['Pięć szczebli kontroli nad dostawcą (Master 5.4 §7.3)'],
 ['Szczebel', 'Co znaczy', 'Kontrola', 'Rola wg MDR'],
]
for s, v in sorted(SZCZEBEL.items()):
    info.append([s, v[0], v[1], v[2]])
info += [
 [],
 ['Pięć klas komponentów wg Master 5.4 §4.2 — składowe I–IV'],
 ['Klasa', 'Nazwa', 'Zakres', 'Co konkretnie w Eternal'],
]
for s, v in SKLADOWE.items():
    info.append([s, v[0], v[1], v[2]])
info += [
 [],
 ['UWAGA METODYCZNA'],
 ['Pole „klasa MDR” w źródłowym rejestrze funkcji (IIA/IIB/III) NIE zostało użyte '
  'do klasyfikacji. Kontrola wykazała, że jest artefaktem ekstrakcji: jako IIb oznaczone '
  'są tam m.in. „Dashboard główny” i „Ręczne dodawanie danych”, które wyrobem nie są. '
  'Warstwa zgodności jest wyprowadzona z definicji Master 5.4 i z treści nazwy funkcji.'],
 ['Kontrola poprawności: korpus wskazuje A3.5, A6.5, A6.8 i D2.x jako warstwę C. '
  'Wszystkie cztery przypadki reguły odtwarzają niezależnie.'],
 ['Kontrola % liczona jawnym wzorem: 0,40 × (szczebel/5) + 0,25 × dane + 0,20 × '
  'wymienialność + 0,15 × wniosek. Wagi są arbitralne, ale jawne — mówią, co uznajemy '
  'za kontrolę. Zmiana wag zmienia wynik i to jest cecha, nie wada.'],
 ['Pokrycie modułu przez kandydata rynkowego jest oceną autorską opartą na znajomości '
  'rynku, nie pomiarem. Wymaga potwierdzenia w rozmowie z dostawcą.'],
 ['Wartości oznaczone [SZACUNEK] to przeliczenie wolumenu na użytkownika, nie cennik '
  'dostawcy. [BRAK] oznacza pozycję bez danych. Kursy: 1 USD = 4,00 PLN, 1 EUR = 4,30 PLN.'],
]
for r_ in info:
    ws.append(r_)
for w, s in zip('ABCDEF', [22, 40, 30, 26, 10, 80]):
    ws.column_dimensions[w].width = s
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.font = Font(size=9)
        if c.column == 1 and c.value and not row[1].value:
            c.font = Font(size=11, bold=True, color='1B3A6B')

del wb['Sheet']
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print('%s -> %d B, arkuszy %d, funkcji %d, klas %d'
      % (OUT, os.path.getsize(OUT), len(wb.sheetnames), len(R), len(K)))
print('warstwy:', dict(c_w), '| certyfikacja wymagana:', c_w.get('C', 0))
