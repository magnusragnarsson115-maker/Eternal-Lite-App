# -*- coding: utf-8 -*-
"""Rejestr 337 funkcji + produkty + branze -> XLSX."""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import rejestr as RJ, dane_produkty as PR, karty

HDR = Font(bold=True, color='FFFFFF'); FILL = PatternFill('solid', fgColor='1F3864')
def sheet(wb, name, rows, widths):
    ws = wb.create_sheet(name)
    for r in rows: ws.append(r)
    for c in ws[1]:
        c.font = HDR; c.fill = FILL; c.alignment = Alignment(vertical='center', wrap_text=True)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    ws.freeze_panes = 'A2'
    for row in ws.iter_rows(min_row=2):
        for c in row: c.alignment = Alignment(vertical='top', wrap_text=True)
    return ws

wb = Workbook(); wb.remove(wb.active)
PRODF = karty.PRODUKT_FUNKCJI

rows = [["Kod", "Nazwa", "Produkt", "Moduł", "Etap", "Priorytet", "Warstwa",
         "Wyrób medyczny", "Klasa MDR", "Klasa komponentu", "Kanał monetyzacji",
         "Waga user", "Waga ekosystem", "Na start", "Próg wyjścia na własne",
         "Kontrola", "W produkcie"]]
for kod in sorted(RJ.R, key=lambda x: (RJ.R[x]['produkt'], RJ.R[x]['modul_kod'], x)):
    v = RJ.R[kod]
    rows.append([kod, v['nazwa'], v['produkt'], v['modul'], v['etap'],
                 karty.karta(kod)['priorytet'], v['warstwa'], v['medical_device'],
                 v['klasa_mdr'], "%s %s" % (v['klasa_komp'], v['klasa_komp_nazwa']),
                 v['kanal'], v['waga_user'], v['waga_eko'], v['dostawca_start'],
                 v['prog_zmiany'], v['kontrola'], "; ".join(PRODF.get(kod, []))])
sheet(wb, 'Rejestr funkcji', rows,
      [9, 42, 18, 26, 8, 9, 8, 12, 10, 30, 34, 11, 13, 40, 34, 30, 26])

rows = [["Produkt", "Claim", "Funkcje", "Odbiorca", "Monetyzacja", "Warstwa", "Etap"]]
for p in PR.PRODUKTY:
    rows.append([p['nazwa'], p['claim'], " · ".join(p['funkcje']), p['odbiorca'],
                 p['monetyzacja'], p['warstwa'], p['etap']])
sheet(wb, 'Produkty', rows, [20, 50, 26, 44, 50, 9, 30])

sheet(wb, 'Branże i nisze', PR.BRANZE, [26, 26, 46, 26, 26])
sheet(wb, 'Monetyzacja', PR.MONETYZACJA, [20, 34, 26, 24, 30])
sheet(wb, 'Build czy buy', PR.ALTERNATYWY, [30, 40, 16, 44])

out = '/home/user/Eternal-Lite-App/out/ETERNAL_REJESTR_FUNKCJI.xlsx'
wb.save(out)
print('%s -> %d B, arkuszy: %d' % (out, os.path.getsize(out), len(wb.sheetnames)))
