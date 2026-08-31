# -*- coding: utf-8 -*-
"""Odpowiedniki rynkowe: 337 funkcji, 22 pozycje dostawcow, 8 agregatorow."""
import json
import os
import sys
import collections

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dane_komponenty import K
from dane_moduly import M as MOD
from dane_rynek import POZYCJE, AGREGATORY, A1_FUNKCJE, TEST

R = json.load(open('build/KOMPONENTY.json'))
OUT = '/home/user/Eternal-Lite-App/out/ETERNAL_ODPOWIEDNIKI_RYNKOWE.xlsx'

GRAN = PatternFill('solid', fgColor='1B3A6B')
RDZA = PatternFill('solid', fgColor='B8431F')
ZIEL = PatternFill('solid', fgColor='D7F0DD')
CZERW = PatternFill('solid', fgColor='F8D7DA')
ZOLT = PatternFill('solid', fgColor='F3E3C3')
CIENKA = Border(*[Side('thin', color='D8D4CE')] * 4)
wb = Workbook()


def arkusz(nazwa, naglowki, wiersze, szer, fill=GRAN):
    ws = wb.create_sheet(nazwa)
    ws.append(naglowki)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = fill
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    for w in wiersze:
        ws.append(w)
    for i, s in enumerate(szer, 1):
        ws.column_dimensions[get_column_letter(i)].width = s
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.font = Font(size=9)
            c.border = CIENKA
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    return ws


# --- 1. Odpowiednik rynkowy per funkcja -----------------------------------
A1MAP = {x[0]: x for x in A1_FUNKCJE}
MODMAP = {k: v for k, v in MOD.items()}


def kupic(r):
    m = MODMAP.get(r['modul'])
    if not m:
        return '—', '—'
    if m[7] >= 80:
        return 'TAK — cały moduł', m[6]
    if m[7] >= 30:
        return 'CZĘŚCIOWO — komponent', m[6]
    return 'NIE — brak kandydata', m[6]


wiersze = []
for r in sorted(R, key=lambda x: (x['modul'][0], int(x['modul'][1:]), x['kod'])):
    kl = K[r['klasa']]
    a1 = A1MAP.get(r['kod'])
    da, kand = kupic(r)
    wiersze.append([
        r['kod'], r['nazwa'], '%s — %s' % (r['modul'], r['modul_nazwa']),
        '%s — %s' % (r['klasa'], kl[0]),
        a1[2] if a1 else kl[0],
        a1[3] if a1 else kl[2],
        a1[4] if a1 else ('TAK — to jest gotowy produkt' if kl[2] and 'BRAK' not in kl[2]
                          else 'NIE — brak dostawcy'),
        kl[1], kand, da, r['warstwa'], r['etap'],
        a1[5] if a1 else kl[6]])
arkusz('Odpowiedniki per funkcja',
 ['Kod', 'Funkcja', 'Moduł', 'Klasa komponentu', 'Odpowiednik rynkowy',
  'Kto to robi (konkurencja / dostawcy)', 'Da się kupić?', 'Alternatywa open source',
  'Kandydat na cały moduł', 'Czy cały moduł do kupienia', 'Warstwa', 'Etap',
  'Nasza decyzja'],
 wiersze, [8, 42, 30, 32, 34, 46, 26, 46, 46, 22, 7, 8, 44])
ws = wb['Odpowiedniki per funkcja']
for row in ws.iter_rows(min_row=2):
    v = str(row[9].value or '')
    row[9].fill = (ZIEL if v.startswith('TAK') else
                   ZOLT if v.startswith('CZĘŚ') else CZERW)

# --- 2. Modul A1 wzorcowy --------------------------------------------------
arkusz('A1 — wzorzec',
 ['Kod', 'Funkcja', 'Odpowiednik rynkowy', 'Kto to robi', 'Da się kupić?',
  'Nasza decyzja'],
 [list(x) for x in A1_FUNKCJE], [8, 46, 30, 40, 50, 60], fill=RDZA)

# --- 3. Pokrycie A1 przez kandydatow --------------------------------------
arkusz('A1 — pokrycie kandydatów',
 ['Kandydat', 'A1.1 sync', 'A1.2 open', 'A1.3 ręcznie', 'A1.4 korekta', 'A1.5 FHIR',
  'A1.6 import', 'A1.7 dedup', 'A1.8 SDK OS', 'A1.9 Station', 'A1.10 storage', 'Pokrycie'],
 [['Vitalera', '●', '○', '—', '—', '●', '○', '○', '●', '—', '○', '85%'],
  ['Terra / Rook / Junction', '●', '—', '—', '—', '○', '—', '○', '●', '—', '—', '70%'],
  ['Thryve (mio)', '●', '—', '—', '—', '●', '—', '○', '●', '—', '—', '70%'],
  ['HealthKit + Health Connect', '○', '—', '—', '—', '—', '—', '—', '●', '—', '○', '70%'],
  ['Medplum (serwer FHIR)', '—', '—', '○', '○', '●', '●', '—', '—', '—', '●', '55%'],
  ['Własne adaptery GATT', '●', '●', '—', '—', '—', '—', '—', '—', '●', '—', '60%'],
  ['LEGENDA', '● pełne', '○ częściowe', '— brak', '', '', '', '', '', '', '', '']],
 [26] + [11] * 10 + [10])

# --- 4. Agregatory pogrupowane --------------------------------------------
arkusz('Agregatory — 8 pozycji',
 ['Dostawca', 'Grupa', 'Klasa', 'Model rozliczenia', 'Pokrycie A1 %',
  'Zgodność z Eternal %', 'Rozwój', 'Adaptowalność', 'Czym się różni'],
 [[n, g, k, m, p, z, r, a, d] for n, g, k, m, p, z, r, a, d in AGREGATORY],
 [30, 30, 18, 30, 12, 14, 40, 34, 60])
ws = wb['Agregatory — 8 pozycji']
for row in ws.iter_rows(min_row=2):
    v = str(row[2].value or '')
    row[2].fill = CZERW if v.startswith('medyczny') else ZIEL

# --- 5. Macierz dostawcow: test otwartego standardu -----------------------
arkusz('Test otwartego standardu',
 ['Pozycja', 'Konkurencja / odpowiednik', '3 opcje rynkowe', '3 z white label / OEM',
  'Wyjście, gdy cała trójka odpadnie', 'Otwarty standard?', 'Rozwój',
  'Wolno budować rdzeń?'],
 [[p[0], p[1], ' · '.join(p[2]), ' · '.join(p[3]), p[4], p[5], p[6], p[7]]
  for p in POZYCJE], [30, 30, 42, 34, 54, 22, 40, 16], fill=RDZA)
ws = wb['Test otwartego standardu']
for row in ws.iter_rows(min_row=2):
    row[7].fill = ZIEL if str(row[7].value) == 'TAK' else CZERW
    row[7].font = Font(size=9, bold=True)

# --- 6. Legenda ------------------------------------------------------------
ws = wb.create_sheet('Legenda', 0)
c_da = collections.Counter(w[9] for w in wiersze)
std = sum(1 for p in POZYCJE if p[7] == 'TAK')
info = [
 ['ETERNAL LIFE — ODPOWIEDNIKI RYNKOWE'],
 [],
 ['Pytanie nadrzędne', TEST[0]],
 ['Odpowiedź TAK', TEST[1]],
 ['Odpowiedź NIE', TEST[2]],
 [],
 ['Wynik testu na 22 pozycjach macierzy dostawców'],
 ['Wolno budować rdzeń', str(std), 'Standard publiczny — mogę odejść, bo mogę to napisać'],
 ['Nie wolno budować rdzenia', str(len(POZYCJE) - std),
  'CGM, baza leków, P1, płatności, smart clothes, CDMO, nanotech — mogą być funkcjami'],
 [],
 ['Czy cały moduł da się kupić — rozkład 337 funkcji'],
] + [[k, str(v)] for k, v in c_da.most_common()] + [
 [],
 ['Arkusze'],
 ['Odpowiedniki per funkcja', 'Każda z %d funkcji z odpowiednikiem rynkowym, listą '
  'dostawców, alternatywą open source i informacją, czy cały moduł da się kupić.'
  % len(R)],
 ['A1 — wzorzec', 'Moduł A1 rozpisany funkcja po funkcji jako wzorzec dla pozostałych.'],
 ['A1 — pokrycie kandydatów', 'Który kandydat pokrywa które funkcje A1 i w jakim stopniu.'],
 ['Agregatory — 8 pozycji', 'Terra, Rook, Junction, Vitalera, Validic, Thryve, SDK '
  'systemowe i własne GATT — pogrupowane, z pokryciem, zgodnością i adaptowalnością.'],
 ['Test otwartego standardu', '22 pozycje z macierzy dostawców: konkurencja, trzy opcje, '
  'white label, wyjście, standard, rozwój.'],
 [],
 ['UWAGI METODYCZNE'],
 ['Macierz dostawców (22 pozycje) pochodzi z korpusu — ETERNAL_Macierz_Dostawcow.xlsx. '
  'Korpus podaje, że standard istnieje w 13 z 22 pozycji; moja klasyfikacja daje %d, '
  'bo pozycje graniczne (LLM z otwartymi wagami, rozpoznawanie posiłków, bazy ćwiczeń) '
  'liczę jako częściowe. Różnica dotyczy klasyfikacji, nie treści.' % std],
 ['Pokrycie funkcji przez kandydata rynkowego jest oceną autorską opartą na publicznej '
  'dokumentacji dostawców, nie pomiarem. Wymaga potwierdzenia w rozmowie z dostawcą.'],
 ['Vitalera nie ma publicznego cennika — pozycja oznaczona [BRAK], nie oszacowana. '
  'Deklaracja CE wg MDR pochodzi od dostawcy i nie została zweryfikowana w EUDAMED.'],
]
for r_ in info:
    ws.append(r_)
for w, s in zip('ABC', [34, 14, 100]):
    ws.column_dimensions[w].width = s
ws['A1'].font = Font(bold=True, size=16, color='B8431F')
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.font = Font(size=9)
        if c.column == 1 and c.value and not row[1].value:
            c.font = Font(size=11, bold=True, color='1B3A6B')

del wb['Sheet']
wb.save(OUT)
print('%s -> %d B, arkuszy %d, funkcji %d' % (OUT, os.path.getsize(OUT),
                                              len(wb.sheetnames), len(R)))
print('cały moduł do kupienia:', dict(c_da))
